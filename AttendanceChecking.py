import datetime
import os
import string
import subprocess
import sys
from pathlib import Path
from time import strftime, gmtime

from dateutil import parser
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Font, Side


class AttendanceChecking:
    """
    * Author: Phuong Le
    * Date: 6 Sep 2018
    * Creating new or modify existing AttendanceChecking.xlsx file with the predefined format
    """

    header_list = ["ID", "Last Name", "First Name", "Group", "Total"]  # list of headers
    start_table_cell = [5, 1]  # beginning cell of the table

    def __init__(self, file_path):
        if self.check_file_exist(file_path):
            # parameters
            self.file_path = file_path
            self.workbook = load_workbook(file_path)
            self.sheet = self.workbook.active

            if (
                    not hasattr(self, "sheet")
                    or (not self.if_standard_excel())
            ):
                print("Invalid file format!")

    @staticmethod
    def check_file_exist(file_path):
        return Path(file_path).is_file()

    @staticmethod
    def new_standard_file(file_path):
        """
        Generate new excel file with pre-defined contents

        :param file_path: path to save new file
        :return: path of generated file
        """
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.merge_cells("A1:D1")
        new_sheet.merge_cells("A2:D2")
        new_sheet.cell(1, 1).value = "DANH SÁCH SINH VIÊN"
        new_sheet.cell(2, 1).value = "YOUR COURSE/SUBJECT/TITLE"
        new_sheet.cell(1, 1).font = AttendanceChecking.cell_format("header")
        new_sheet.cell(2, 1).font = AttendanceChecking.cell_format("header")
        new_sheet.cell(3, 4).value = "1 = prensent"
        new_sheet.cell(3, 4).font = AttendanceChecking.cell_format("header")
        new_sheet.cell(4, 4).value = "blank = absent"
        new_sheet.cell(4, 4).font = AttendanceChecking.cell_format("header")

        for col in range(
                AttendanceChecking.start_table_cell[1],
                len(AttendanceChecking.header_list) + 1,
        ):
            new_sheet.cell(
                AttendanceChecking.start_table_cell[0], col
            ).value = AttendanceChecking.header_list[
                col - AttendanceChecking.start_table_cell[1]
                ]
            AttendanceChecking.format_font(
                new_sheet, "header", [AttendanceChecking.start_table_cell[0], col]
            )
            new_sheet.cell(
                AttendanceChecking.start_table_cell[0], col
            ).border = AttendanceChecking.cell_format("thin_border")

        # format table
        AttendanceChecking.page_setups(new_sheet)

        new_wb.save(file_path)
        return file_path

    @staticmethod
    def page_setups(sheet):
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToHeight = 1
        sheet.page_setup.fitToWidth = 10

    @staticmethod
    def get_cell_index(sheet, value):
        """
        :param sheet:
        :param value:
        :return: [row, col] or -1 if cell not exist
        """
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row, col).value == value:
                    return [row, col]
        return -1

    @staticmethod
    def cell_format(type):
        header = Font(
            name="Times New Roman",
            size=13,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="000000",
        )

        normal = Font(
            name="Times New Roman",
            size=13,
            bold=False,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="000000",
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        return {"header": header, "normal": normal, "thin_border": thin_border}.get(
            type, normal
        )

    @staticmethod
    def format_font(sheet, type, cell):
        sheet.cell(cell[0], cell[1]).font = AttendanceChecking.cell_format(type)

    @staticmethod
    def format_border(sheet, type, cell):
        sheet.cell(cell[0], cell[1]).border = AttendanceChecking.cell_format(type)

    def get_begin_cell_index(self, value):
        return self.get_cell_index(self.sheet, value)

    def get_end_cell_index(self, value):
        return self.get_cell_index(self.sheet, value)

    def validate_table(self, end_col_header):
        if self.sheet.max_column != self.get_end_cell_index(end_col_header)[1]:
            return -1
        else:
            return 0

    def if_standard_excel(self):
        # check if having any sheets
        if not hasattr(self, "sheet"):
            return False

        try:
            # Column ID
            id = self.get_cell_index(self.sheet, "ID")

            # Column Last Name
            last_name = self.get_cell_index(self.sheet, "Last Name")

            # Column First Name
            first_name = self.get_cell_index(self.sheet, "First Name")

            # Column Group
            group = self.get_cell_index(self.sheet, "Group")

            # column Total
            total = self.get_cell_index(self.sheet, "Total")

            # column next to one Total
            next_total = self.sheet.cell(total[0], total[1] + 1).value

        except TypeError:
            return False

        if (
                (id == -1)
                or (last_name == -1)
                or (first_name == -1)
                or (group == -1)
                or (total == -1)
                # No more column allowed next to one Total
                or (next_total is not None)
        ):
            return False
        else:
            for col in range(group[1] + 1, total[1]):
                cell_type = type(self.sheet.cell(id[0], col).value)
                if cell_type == str:
                    try:
                        cell_type = type(
                            parser.parse(self.sheet.cell(id[0], col).value)
                        )
                        if cell_type != datetime.datetime:
                            print(
                                "Cell must be a date: "
                                + str([id[0], col])
                                + " - "
                                + str(self.sheet.cell(id[0], col).value)
                                + " is a "
                                + str(cell_type)
                            )
                            return False
                    except ValueError:
                        print(
                            "Unknown string format: "
                            + str([id[0], col])
                            + " - "
                            + str(self.sheet.cell(id[0], col).value)
                        )
                        return False
                else:
                    if cell_type != datetime.datetime:
                        print(
                            " Cell must be a date: "
                            + str([id[0], col])
                            + " - "
                            + str(self.sheet.cell(id[0], col).value)
                            + " is a "
                            + str(cell_type)
                        )
                        return False
            return True

    def open_active_sheet(self):
        return load_workbook(self.file_path).active

    def shift_col_to_right(self, col, from_row):
        """
        Shift a column to the right one

        :param col: column wanted to be shifted
        :param from_row: where contents of that column starts
        """
        for row in range(from_row, self.sheet.max_row + 1):
            self.sheet.cell(row, col + 1).value = self.sheet.cell(row, col).value
            self.sheet.cell(row, col).value = ""

    def set_header(self, row, col, value):
        self.sheet.cell(row, col).value = value
        self.format_font(self.sheet, "header", [row, col])
        self.sheet.cell(row, col).border = AttendanceChecking.cell_format("thin_border")

    @staticmethod
    def get_col_letter(col):
        """
        Convert column letter from numeric to alphabetical letter

        :param col: column number such as 1, 2, 3
        :return: alphabetical letter such as A, B, C
        """

        title = ""
        alist = string.ascii_uppercase
        while col:
            mod = (col - 1) % 26 #26 alphabetical letters
            col = int((col - mod) / 26)
            title += alist[mod]
        return title[::-1]

    def open_in_system_app(self):
        """
        Call system app to open current excel file

        """

        if sys.platform.startswith("darwin"):
            subprocess.call(("open", self.file_path))
        elif os.name == "nt":  # For Windows
            os.startfile(self.file_path)
        elif os.name == "posix":  # For Linux, Mac, etc.
            subprocess.call(("xdg-open", self.file_path))

    @staticmethod
    def get_current_date():
        """
        Get current system time with pre-defined format

        :return: m/d/Y
        """

        return strftime("%m/%d/%Y", gmtime())

    def update_total_col(
            self, from_row, to_row, from_col, to_col, total_header
    ):
        """
        Automatically update sum function of each cell in column Total, sum of a rect range of cell with 4 coordinates: from_row, to_row, from_col, to_col

        :param from_row: start row of the rect range to sum
        :param to_row: end row of the rect range to sum
        :param from_col: start column of the rect range to sum
        :param to_col: end column of the rect range to sum
        :param total_header: [row column] of cell with content Total
        """
        for row in range(from_row, to_row + 1):
            sum_func = (
                    "=COUNTBLANK("
                    + str(self.get_col_letter(from_col))
                    + str(row)
                    + ":"
                    + str(self.get_col_letter(to_col))
                    + str(row)
                    + ")"
            )
            self.sheet.cell(row, total_header[1]).value = sum_func

    def get_total_absence(self, mssv):
        """
        Get the number of absence of a specific student based on his ID

        :return: None if not standard excel content
        :return: number of absence
        :param mssv: which needed to get number of absence
        """
        if not self.if_standard_excel():
            return None
        header_total = self.get_end_cell_index("Total")
        header_group = self.get_cell_index(self.sheet, "Group")
        mssv_cell = self.get_cell_index(self.sheet, mssv)
        total_absence = 0
        for col in range(header_group[1] + 1, header_total[1]):
            if self.sheet.cell(mssv_cell[0], col).value is None:
                total_absence += 1
        return total_absence

    def get_total_mssv(self):
        """
        Get the number of students in all

        :return: total of students
        """
        if not self.if_standard_excel():
            return None
        total_mssv = 0
        header_id = self.get_cell_index(self.sheet, "ID")
        for row in range(header_id[0] + 1, self.sheet.max_row + 1):
            if self.sheet.cell(row, header_id[1]).value is not None:
                total_mssv += 1
        return total_mssv

    def if_exist_table_header(self, value, cell):
        if not self.if_standard_excel():
            return False

        if self.sheet.cell(cell[0], cell[1]).value == value:
            return True
        else:
            return False

    def start_checking(self, mssv_array):
        """
        Execute attendance checking for each id in ID column

        :param mssv_array: array of intended checking mssv (ID):
        :return: failed_case - list of not checked id yet
        """
        if not self.if_standard_excel():
            print("Invalid file format!")
            # print('Index of "Total": ' + str(self.get_cell_index(self.sheet, "Total")))
            # print("Find max column is: " + str(self.sheet.max_column))
            return mssv_array  # wrong form

        failed_case = []  # Fail to mark as present
        col_total = self.get_cell_index(self.sheet, "Total")
        new_checking = col_total
        group = self.get_cell_index(self.sheet, "Group")
        header_row = self.get_begin_cell_index("ID")[0]
        id_col = self.get_begin_cell_index("ID")[1]

        if not self.if_exist_table_header(
                self.get_current_date(), [col_total[0], col_total[1] - 1]
        ):
            # Shift col with header Total to one more col right
            self.shift_col_to_right(col_total[1], col_total[0])

            # Insert new day header
            self.set_header(col_total[0], col_total[1], self.get_current_date())

            # update indice for cell with header Total
            col_total = [col_total[0], col_total[1] + 1]

            # format cell
            self.format_font(self.sheet, "header", [col_total[0], col_total[1]])
            self.format_border(self.sheet, "thin_border", [col_total[0], col_total[1]])
        else:
            new_checking = [col_total[0], col_total[1] - 1]

        # update index for header Total
        col_total[0] = new_checking[0]
        col_total[1] = new_checking[1] + 1

        # start mark attendace
        checked = False
        for id in mssv_array:
            for row in range(header_row + 1, self.sheet.max_row + 1):
                if id == self.sheet.cell(row, id_col).value:
                    # presence
                    self.sheet.cell(row, new_checking[1]).value = 1
                    checked = True
                    break
            if not checked:
                failed_case.append(id)

        # update Count function whether any data or not
        self.update_total_col(
            col_total[0] + 1,
            self.sheet.max_row,
            group[1] + 1,
            new_checking[1],
            col_total
        )
        # set some formats
        AttendanceChecking.page_setups(self.sheet)
        self.workbook.save(self.file_path)
        return failed_case
